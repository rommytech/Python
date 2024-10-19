import xlsxwriter
import os
import openpyxl  # Corregido: Importación correcta de openpyxl

# -*- coding: utf-8 -*-
from .cuenta_galactica import CuentaGalactica

class Personaje:
    def __init__(self, nombre):
        self.nombre = nombre
        self.cuenta = CuentaGalactica()
        self.crear_excel_v2()  # Corregido: Llamada al método, no a la función en sí
    
    def crear_excel_v1(self):
        name_file = self.nombre + ".xlsx"
        if os.path.exists(name_file):
            pass  # No hacer nada si el archivo ya existe
        else:
            # Crear el archivo Excel
            workbook = xlsxwriter.Workbook(name_file)
            worksheet = workbook.add_worksheet()  # Creación de la hoja de trabajo
            worksheet.write("A1", self.nombre)    # Escribir el nombre en la celda A1
            workbook.close()

    def crear_excel_v2(self):
        # Verificar si el archivo ya existe antes de intentar cargarlo
        if not os.path.exists("personajes.xlsx"):
            workbook = openpyxl.Workbook()  # Crear un nuevo workbook si no existe el archivo
        else:
            try:
                workbook = openpyxl.load_workbook("personajes.xlsx")  # Cargar el archivo existente
            except Exception as e:
                workbook = openpyxl.Workbook()  # Crear un nuevo workbook si hay error en la carga
        
        hoja_activa = workbook.active
        ultimo_row = hoja_activa.max_row + 1
        
        celda = "A" + str(ultimo_row)
        hoja_activa[celda] = self.nombre
        workbook.save("personajes.xlsx")
    
    def depositar_creditos(self, monto):
        self.cuenta.depositar(monto)

    def retirar_creditos(self, monto):
        self.cuenta.retirar(monto)

    def mostrar_saldo(self):
        return self.cuenta.saldo
